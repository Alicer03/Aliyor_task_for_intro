import datetime
import pandas as pd

import openpyxl as op
import telebot

bot = telebot.TeleBot('6960456784:AAE7NH4bqXQ97DIpzwsmk3SoYs8XBmbFMog')

filename = 'data_id.xlsx'

wb = op.load_workbook(filename, data_only=True)
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

max_rows = sheet1.max_row
max_columns = sheet1.max_column

# the dicts
students = {}

for i in range(4, max_rows + 1):
    student_id = sheet1.cell(row=i, column=4).value
    students[student_id] = {}

    for j in range(5, max_columns + 1):
        day = sheet1.cell(row=2, column=j).value
        room = sheet1.cell(row=i, column=j).value
        day_tag = sheet1.cell(row=3, column=j).value
        time = sheet1.cell(row=1, column=j).value

        if room is not None:
            if day not in students[student_id]:
                students[student_id][day] = {}

            if room not in students[student_id][day] or room in students[student_id][day]:
                students[student_id][day][day_tag] = []
                students[student_id][day][day_tag].append(room)
                students[student_id][day][day_tag].append(time)

                for h in range(1, max_rows + 1):
                    subject = sheet2.cell(row=h, column=4).value

                    if subject is not None:
                        if day_tag == subject:

                            for e in range(1, max_columns + 1):
                                subject = sheet2.cell(row=h, column=e).value
                                if room == sheet2.cell(row=3, column=e).value:
                                    students[student_id][day][day_tag].append(subject)


@bot.message_handler(commands=['start'])
def main(message):
    bot.send_message(message.chat.id, "Write your id ")


@bot.message_handler()
def info(message):
    student = message.text
    for n in students[student]:
        count = " "
        for k in students[student][n]:
            a = pd.DataFrame(students[student][n][k], columns=[f'{n} {count}'])
            a.index = ["Room: ", 'Time: ', 'Subject: ']
            bot.send_message(message.chat.id, f'{a}')
            current_time = datetime.datetime.now().strftime("%H:%M:%S")
            bot.send_message(message.chat.id, f"Tekushchee vremya: {current_time}")
            count = 0
            count =+ 2

    # b = pd.DataFrame(students[student]['Tuesday'])
    # b.index = ["Room: ", 'Time: ', 'Subject: ']
    # c = pd.DataFrame(students[student]['Wednesday'])
    # c.index = ["Room: ", 'Time: ', 'Subject: ']
    # d = pd.DataFrame(students[student]['Thursday'])
    # d.index = ["Room: ", 'Time: ', 'Subject: ']
    # f = pd.DataFrame(students[student]['Friday'])
    # f.index = ["Room: ", 'Time: ', 'Subject: ']


    # bot.send_message(message.chat.id, f'{b}')
    # bot.send_message(message.chat.id, f'{c}')
    # bot.send_message(message.chat.id, f'{d}')
    # bot.send_message(message.chat.id, f'{f}')


bot.infinity_polling()
